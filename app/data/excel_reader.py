import openpyxl


class ExcelReader:
    def read_data(self, file_path):
        """
        读取Excel数据文件
        :param file_path: Excel文件路径
        :return: 读取的数据
        """
        try:
            # 使用openpyxl加载Excel文件
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            
            # 获取最大行数
            max_row = sheet.max_row
            
            if max_row < 2:
                raise Exception("Excel文件中没有足够的数据行,或检查表格中标题行是否存在()")
            
            # 读取Excel文件的第一行作为标题
            headers = [cell.value for cell in sheet[1]]
            
            # 获取列索引
            try:
                mac_start_idx = headers.index("MacStart") + 1
                mac_end_idx = headers.index("MacEnd") + 1
                gpon_sn_idx = headers.index("GPON-SN") + 1
                device_id_idx = headers.index("设备标识") + 1
                model_idx = headers.index("设备型号") + 1
            except ValueError as e:
                raise Exception("Excel文件缺少必要的列标题")
            
            # 获取第一行和最后一行数据
            first_row = next(sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=len(headers), values_only=True))
            last_row = next(sheet.iter_rows(min_row=max_row, max_row=max_row, min_col=1, max_col=len(headers), values_only=True))
            
            fields = [(first_row[mac_start_idx - 1], "mac_begin"), (first_row[gpon_sn_idx - 1], "sn_begin"), 
                      (first_row[device_id_idx - 1], "devid_start"), (first_row[model_idx - 1], "h_model"), 
                      (last_row[mac_end_idx - 1], "mac_end"), (last_row[gpon_sn_idx - 1], "sn_end"), 
                      (last_row[device_id_idx - 1], "devid_end")]
            
            # 检查数据中是否有空格
            for value, attr in fields:
                if " " in value:
                    raise Exception("源文件数据存在空格，请检查。")

            workbook.close()
            
            # 获取有效数据行数（减去标题行）
            dev_num = max_row - 1
            
            # 构建返回数据
            data = {
                "mac_begin": fields[0][0],
                "sn_begin": fields[1][0],
                "devid_start": fields[2][0],
                "h_model": fields[3][0],
                "mac_end": fields[4][0],
                "sn_end": fields[5][0],
                "devid_end": fields[6][0],
                "dev_num": dev_num
            }
            
            return data

        except Exception as e:
            raise Exception(f"读取Excel文件时发生错误: {str(e)}")
