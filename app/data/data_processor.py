import openpyxl


class DataProcessor:
    def load_basecfg_data(self, basecfg_path, market_value, model_value, dev_type):
        """
        加载基础配置数据
        :param basecfg_path: 基础配置文件路径
        :param market_value: 市场值
        :param model_value: 机型值
        :param dev_type: 设备类型
        :return: 基础配置数据和表头
        """
        try:
            if not openpyxl.load_workbook(basecfg_path):
                return [], []

            workbook = openpyxl.load_workbook(basecfg_path)
            sheet = workbook.active
            basecfg_headers = [cell.value for cell in sheet[2]]  # 获取表头信息
            
            # 从Excel表格中获取数据，并创建一个字典数据，查找数据的关键字是市场和机型
            basecfg_data = []
            for row in sheet.iter_rows(min_row=3, values_only=False):
                if row[basecfg_headers.index("ponMARKET")].value == market_value and row[basecfg_headers.index("h_model")].value == model_value and row[basecfg_headers.index("DEVTYPE")].value == dev_type:
                    row_data = {basecfg_headers[i]: (cell.value if cell is not None and hasattr(cell, 'value') else "") for i, cell in enumerate(row)}
                    basecfg_data.append(row_data)

            return basecfg_data, basecfg_headers

        except Exception as e:
            raise Exception(f"加载基础配置文件失败: {e}")

    def process_xml_data(self, data):
        """
        处理XML数据
        :param data: XML数据
        :return: 处理后的数据
        """
        # 这里可以添加数据处理逻辑
        return data

    def process_excel_data(self, data):
        """
        处理Excel数据
        :param data: Excel数据
        :return: 处理后的数据
        """
        # 这里可以添加数据处理逻辑
        return data
